
# app_v09.py
# EFESO FKA – v09 (improved)
import streamlit as st
import pandas as pd
import numpy as np
import altair as alt
from io import BytesIO
from openpyxl import load_workbook

st.set_page_config(page_title="FKA v09 – EFESO", layout="wide")
PRIMARY = "#f17726"
BLUE    = "#2b6cb0"
LIGHT   = "#90caf9"
RED     = "#c0392b"

with st.sidebar:
    st.markdown("### Anzeige-Optionen")
    BAR_SIZE_MAIN = st.slider("Balkenbreite – Hauptdiagramme", 10, 50, 24, 2)
    BAR_SIZE_SMALL = st.slider("Balkenbreite – Drilldowns", 8, 40, 18, 2)
    LABEL_ANGLE = st.slider("X-Achse Winkel", 0, 90, 0, 5)
    st.markdown("---")
    st.markdown("##### Optionales Logo")
    logo = st.file_uploader("Logo (PNG/JPG, optional)", type=["png", "jpg", "jpeg"])

col1, col2 = st.columns([1,5])
with col1:
    if logo:
        st.image(logo, use_column_width=True)
    else:
        st.markdown(
            f"""
            <svg width="130" height="36" viewBox="0 0 260 72">
              <rect width="260" height="72" rx="8" fill="{PRIMARY}" opacity="0.15"></rect>
              <text x="14" y="46" style="font: 700 26px sans-serif; fill:{PRIMARY};">EFESO</text>
            </svg>
            """,
            unsafe_allow_html=True,
        )
with col2:
    st.title("FKA v09 – EFESO")
    st.caption("Funktionsmatrix · Kosten · Technik · Abweichungen")

START_COL = 8
ROW_H1   = 0
ROW_H2   = 1
ROW_H1_W = 3
ROW_H2_W = 4
ROW_H1_C = 6
ROW_H2_C = 7

TECH_SHEET = "SLAVE_Techn.Bewertung"
COST_SHEET = "SLAVE_Funktions-Kostenstruktur"

def sanitize_name(name: str) -> str:
    return str(name).replace(" ", "_").replace("-", "_").replace(".", "_")

def read_sheet_values(file_bytes: bytes, sheet_name: str):
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    if sheet_name not in wb.sheetnames:
        return None
    ws = wb[sheet_name]
    values = []
    for row in ws.iter_rows(values_only=True):
        values.append(list(row))
    return values

def to_float(x):
    if x is None or str(x).strip() == "":
        return np.nan
    try:
        return float(str(x).replace(",", "."))
    except Exception:
        return np.nan

def parse_cost_structure(file_bytes: bytes):
    values = read_sheet_values(file_bytes, COST_SHEET)
    if values is None:
        raise ValueError(f"Blatt '{COST_SHEET}' nicht gefunden.")
    max_cols = max(len(r) for r in values if r)
    for _ in range(max(ROW_H2_C+1 - len(values), 0)):
        values.append([None]*max_cols)
    def row_vals(row_idx):
        row = values[row_idx] if row_idx < len(values) else []
        return list(row[START_COL:]) if len(row)>START_COL else []
    h1_names = row_vals(ROW_H1)
    h2_names = row_vals(ROW_H2)
    h1_w     = row_vals(ROW_H1_W)
    h2_w     = row_vals(ROW_H2_W)
    h1_c     = row_vals(ROW_H1_C)
    h2_c     = row_vals(ROW_H2_C)
    records_h2 = []
    last_h1 = None
    for idx in range(max(len(h2_names), len(h1_names))):
        h1 = h1_names[idx] if idx < len(h1_names) else None
        h2 = h2_names[idx] if idx < len(h2_names) else None
        w2 = h2_w[idx] if idx < len(h2_w) else None
        c2 = h2_c[idx] if idx < len(h2_c) else None
        if h1 is not None and str(h1).strip():
            last_h1 = str(h1).strip()
        if h2 is None or str(h2).strip() == "":
            continue
        records_h2.append({
            "Hauptfunktion": last_h1,
            "Nebenfunktion": str(h2).strip(),
            "H2_Gewicht_%": to_float(w2),
            "H2_Kosten": to_float(c2)
        })
    h1_records = []
    seen = set()
    for idx, val in enumerate(h1_names):
        if val is None or str(val).strip() == "":
            continue
        h1 = str(val).strip()
        if h1 in seen: continue
        seen.add(h1)
        w = h1_w[idx] if idx < len(h1_w) else None
        c = h1_c[idx] if idx < len(h1_c) else None
        h1_records.append({"Hauptfunktion": h1, "H1_Gewicht_%": to_float(w), "H1_Kosten": to_float(c)})
    df_h2 = pd.DataFrame(records_h2).dropna(subset=["Nebenfunktion"])
    df_h1 = pd.DataFrame(h1_records).dropna(subset=["Hauptfunktion"])
    if "Hauptfunktion" in df_h2:
        df_h2["Hauptfunktion"] = df_h2["Hauptfunktion"].ffill()
        df_h2 = df_h2.dropna(subset=["Hauptfunktion"])
    return df_h1, df_h2

def parse_tech(file_bytes: bytes, df_h2: pd.DataFrame):
    values = read_sheet_values(file_bytes, TECH_SHEET)
    if values is None:
        return (pd.DataFrame(columns=["Nebenfunktion","TechScore"]),
                pd.DataFrame(columns=["Hauptfunktion","TechScore_mean","TechScore_weighted"]))
    names, scores = [], []
    for r in values[5:400]:
        if r is None: continue
        name = r[1] if len(r)>1 else None
        score = r[17] if len(r)>17 else None
        if name is None or str(name).strip()=="": continue
        s = to_float(score)
        if pd.isna(s): continue
        names.append(str(name).strip()); scores.append(s)
    df_h2_tech = pd.DataFrame({"Nebenfunktion": names, "TechScore": scores})
    df_map = df_h2[["Nebenfunktion","Hauptfunktion","H2_Gewicht_%"]].drop_duplicates()
    tech_join = df_h2_tech.merge(df_map, on="Nebenfunktion", how="left")
    agg_mean = tech_join.groupby("Hauptfunktion", dropna=True)["TechScore"].mean().reset_index().rename(columns={"TechScore":"TechScore_mean"})
    tech_join["w"] = tech_join["H2_Gewicht_%"].astype(float)/100.0
    tech_join["wx"] = tech_join["w"]*tech_join["TechScore"]
    agg_w = tech_join.groupby("Hauptfunktion", dropna=True).agg({"wx":"sum","w":"sum"}).reset_index()
    agg_w["TechScore_weighted"] = np.where(agg_w["w"]>0, agg_w["wx"]/agg_w["w"], np.nan)
    tech_h1 = agg_mean.merge(agg_w[["Hauptfunktion","TechScore_weighted"]], on="Hauptfunktion", how="outer")
    return df_h2_tech, tech_h1

def collect_product(file, raw_name):
    file_bytes = file.getvalue() if hasattr(file, "getvalue") else file.read()
    name = sanitize_name(raw_name)
    try:
        df_h1, df_h2 = parse_cost_structure(file_bytes)
    except Exception as e:
        st.error(f"[{name}] Kostenstruktur konnte nicht gelesen werden: {e}")
        return None
    df_h2_tech, df_h1_tech = parse_tech(file_bytes, df_h2)
    return {"name": name, "h1": df_h1, "h2": df_h2, "tech_h2": df_h2_tech, "tech_h1": df_h1_tech}

uploaded = st.file_uploader("Excel-Dateien (.xlsx/.xlsm) – je Produkt eine Datei", type=["xlsx","xlsm"], accept_multiple_files=True)
if not uploaded:
    st.info("Bitte Dateien hochladen.")
    st.stop()

products = []
for f in uploaded:
    p = collect_product(f, f.name.rsplit(".",1)[0])
    if p: products.append(p)
if len(products)==0:
    st.error("Keine gültigen Produkte erkannt."); st.stop()

tabMatrix, tabCosts, tabTech, tabDiff = st.tabs(["Funktionsmatrix","Funktionenkosten","Technik Bewertung","Top Kostenabweichung"])

with tabMatrix:
    st.subheader("Funktionsmatrix")
    pname = st.selectbox("Produkt wählen", [p["name"] for p in products], key="mx_prod")
    P = next(p for p in products if p["name"]==pname)
    mat = P["h2"][["Hauptfunktion","Nebenfunktion","H2_Gewicht_%"]]
    if mat.empty:
        st.warning("Keine H2-Daten gefunden.")
    else:
        h1_order = list(P["h1"]["Hauptfunktion"])
        h2_order = mat["Nebenfunktion"].tolist()
        heat = alt.Chart(mat).mark_rect().encode(
            x=alt.X("Hauptfunktion:N", sort=h1_order, title="Hauptfunktion", axis=alt.Axis(labelAngle=LABEL_ANGLE)),
            y=alt.Y("Nebenfunktion:N", sort=h2_order, title="Nebenfunktion"),
            color=alt.Color("H2_Gewicht_%:Q", scale=alt.Scale(scheme="oranges"), title="Gewicht (%)"),
            tooltip=["Hauptfunktion","Nebenfunktion", alt.Tooltip("H2_Gewicht_%:Q", format=".0f")]
        ).properties(height=540)
        txt = alt.Chart(mat).mark_text(color="black", fontSize=11).encode(
            x=alt.X("Hauptfunktion:N", sort=h1_order),
            y=alt.Y("Nebenfunktion:N", sort=h2_order),
            text=alt.Text("H2_Gewicht_%:Q", format=".0f")
        )
        st.altair_chart(heat+txt, use_container_width=True)

with tabCosts:
    st.subheader("Funktionenkosten")
    pname = st.selectbox("Produkt wählen", [p["name"] for p in products], key="cost_prod")
    P = next(p for p in products if p["name"]==pname)

    left, right = st.columns([1,1])
    with left:
        st.markdown("#### Kosten je Hauptfunktion (Zeile 7)")
        df_h1c = P["h1"][["Hauptfunktion","H1_Kosten"]].dropna()
        if not df_h1c.empty:
            ch = alt.Chart(df_h1c).mark_bar(size=BAR_SIZE_MAIN, color=BLUE).encode(
                x=alt.X("Hauptfunktion:N", sort=list(df_h1c["Hauptfunktion"]), axis=alt.Axis(labelAngle=LABEL_ANGLE)),
                y=alt.Y("H1_Kosten:Q", title="Kosten [€]"),
                tooltip=["Hauptfunktion", alt.Tooltip("H1_Kosten:Q", format=".2f")]
            ).properties(height=260)
            st.altair_chart(ch, use_container_width=True)
    with right:
        st.markdown("#### Kosten je Nebenfunktion (Zeile 8) – Drilldown")
        h1_list = list(P["h1"]["Hauptfunktion"])
        sel_h1 = st.selectbox("Hauptfunktion", h1_list, key="dd_h1")
        df_h2c = P["h2"].query("Hauptfunktion == @sel_h1")[["Nebenfunktion","H2_Kosten"]]
        if not df_h2c.empty:
            ch2 = alt.Chart(df_h2c).mark_bar(size=BAR_SIZE_SMALL, color=LIGHT).encode(
                x=alt.X("Nebenfunktion:N", sort=list(df_h2c["Nebenfunktion"]), axis=alt.Axis(labelAngle=LABEL_ANGLE)),
                y=alt.Y("H2_Kosten:Q", title="Kosten [€]"),
                tooltip=["Nebenfunktion", alt.Tooltip("H2_Kosten:Q", format=".2f")]
            ).properties(height=260)
            st.altair_chart(ch2, use_container_width=True)

with tabTech:
    st.subheader("Technik Bewertung")
    colA, colB = st.columns(2)
    with colA:
        pA_name = st.selectbox("Produkt A", [p["name"] for p in products], key="tA")
    with colB:
        pB_name = st.selectbox("Produkt B", [p["name"] for p in products], key="tB")
    PA = next(p for p in products if p["name"]==pA_name)
    PB = next(p for p in products if p["name"]==pB_name)

    st.markdown("#### Nebenfunktionen – Linienvergleich")
    A = PA["tech_h2"].copy(); A["Produkt"] = pA_name
    B = PB["tech_h2"].copy(); B["Produkt"] = pB_name
    techAB = pd.concat([A,B], ignore_index=True)
    h2_order = techAB["Nebenfunktion"].drop_duplicates().tolist()
    line = alt.Chart(techAB).mark_line(point=alt.OverlayMarkDef(size=40), strokeWidth=3).encode(
        x=alt.X("Nebenfunktion:N", sort=h2_order, axis=alt.Axis(labelAngle=LABEL_ANGLE), title="Nebenfunktion"),
        y=alt.Y("TechScore:Q", title="Technische Bewertung"),
        color=alt.Color("Produkt:N", scale=alt.Scale(range=[BLUE,LIGHT])),
        tooltip=["Produkt","Nebenfunktion", alt.Tooltip("TechScore:Q", format=".2f")]
    ).properties(height=320)
    st.altair_chart(line, use_container_width=True)

    st.markdown("#### Hauptfunktionen – gruppiert (Mean / Weighted)")
    A_h1 = PA["tech_h1"].copy(); A_h1["Produkt"]=pA_name
    B_h1 = PB["tech_h1"].copy(); B_h1["Produkt"]=pB_name
    A_long = A_h1.melt(id_vars=["Hauptfunktion","Produkt"], var_name="Metrik", value_name="TechScore")
    B_long = B_h1.melt(id_vars=["Hauptfunktion","Produkt"], var_name="Metrik", value_name="TechScore")
    H1_all = pd.concat([A_long,B_long], ignore_index=True).dropna(subset=["TechScore"])
    h1_order = list(PA["h1"]["Hauptfunktion"])
    chart = (alt.Chart(H1_all).mark_bar(size=BAR_SIZE_SMALL)
        .encode(
            x=alt.X("Hauptfunktion:N", sort=h1_order, axis=alt.Axis(labelAngle=LABEL_ANGLE)),
            y=alt.Y("TechScore:Q", title="TechScore"),
            color=alt.Color("Produkt:N", scale=alt.Scale(range=[BLUE,LIGHT])),
            column=alt.Column("Metrik:N", title=None, spacing=10),
            xOffset="Produkt:N",
            tooltip=["Produkt","Hauptfunktion","Metrik", alt.Tooltip("TechScore:Q", format=".2f")]
        ).properties(height=280))
    st.altair_chart(chart, use_container_width=True)

with tabDiff:
    st.subheader("Top Kostenabweichung")
    colA, colB = st.columns(2)
    with colA:
        pa = st.selectbox("Produkt A", [p["name"] for p in products], key="dA")
    with colB:
        pb = st.selectbox("Produkt B", [p["name"] for p in products], key="dB")
    PA = next(p for p in products if p["name"]==pa)
    PB = next(p for p in products if p["name"]==pb)
    A = PA["h2"][["Hauptfunktion","Nebenfunktion","H2_Kosten"]].rename(columns={"H2_Kosten":"Cost_A"})
    B = PB["h2"][["Hauptfunktion","Nebenfunktion","H2_Kosten"]].rename(columns={"H2_Kosten":"Cost_B"})
    cmp = A.merge(B, on=["Hauptfunktion","Nebenfunktion"], how="outer")
    cmp["Delta_(B-A)"] = cmp["Cost_B"].fillna(0) - cmp["Cost_A"].fillna(0)
    top = cmp.copy(); top["absDelta"]=top["Delta_(B-A)"].abs()
    top = top.sort_values("absDelta", ascending=False).head(10)
    st.markdown("#### Top 10 Abweichungen")
    st.dataframe(top[["Hauptfunktion","Nebenfunktion","Cost_A","Cost_B","Delta_(B-A)"]], use_container_width=True, hide_index=True)
    if not top.empty:
        bar = alt.Chart(top).mark_bar(size=BAR_SIZE_SMALL).encode(
            x=alt.X("Nebenfunktion:N", sort=list(top["Nebenfunktion"]), axis=alt.Axis(labelAngle=LABEL_ANGLE)),
            y=alt.Y("Delta_(B-A):Q", title="Delta [€] (B - A)"),
            color=alt.condition(alt.datum["Delta_(B-A)"] > 0, alt.value(BLUE), alt.value(RED)),
            tooltip=["Hauptfunktion","Nebenfunktion",
                     alt.Tooltip("Cost_A:Q", format=".2f"),
                     alt.Tooltip("Cost_B:Q", format=".2f"),
                     alt.Tooltip("Delta_(B-A):Q", format=".2f")]
        ).properties(height=280)
        st.altair_chart(bar, use_container_width=True)
